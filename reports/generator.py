"""
Report Generator
Generates PDF and Excel reports for customers
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import logging
import os
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Any
from uuid import UUID
from pathlib import Path
from io import BytesIO

from config.settings import settings
from database.crud import Database, get_db_connection

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generates various reports for customers:
    - Weekly/Monthly alert digest
    - Watchlist status summary
    - Single trademark analysis
    - Full portfolio report
    """
    
    def __init__(self, db: Database = None):
        self.db = db or Database(get_db_connection())
        self.output_dir = Path(settings.paths.report_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_report(
        self,
        user_id: UUID,
        report_type: str,
        parameters: Dict[str, Any] = None
    ) -> Dict:
        """
        Generate a report based on type.
        
        Args:
            user_id: User requesting the report
            report_type: Type of report
            parameters: Additional parameters (date range, watchlist_ids, etc.)
        
        Returns:
            Report metadata including file path
        """
        parameters = parameters or {}
        
        # Create report record
        report_id = self._create_report_record(user_id, report_type, parameters)
        
        try:
            # Update status to generating
            self._update_report_status(report_id, 'generating')
            
            # Generate based on type
            if report_type == 'weekly_digest':
                file_path = self._generate_weekly_digest(user_id, report_id, parameters)
            elif report_type == 'monthly_summary':
                file_path = self._generate_monthly_summary(user_id, report_id, parameters)
            elif report_type == 'watchlist_status':
                file_path = self._generate_watchlist_status(user_id, report_id, parameters)
            elif report_type == 'single_trademark':
                file_path = self._generate_single_trademark_report(
                    user_id, report_id, parameters.get('watchlist_id'), parameters
                )
            elif report_type == 'full_portfolio':
                file_path = self._generate_portfolio_report(user_id, report_id, parameters)
            else:
                raise ValueError(f"Unknown report type: {report_type}")
            
            # Get file size
            file_size = os.path.getsize(file_path) if file_path else 0
            
            # Update status to completed
            self._update_report_status(
                report_id, 'completed',
                file_path=file_path,
                file_size=file_size
            )
            
            return {
                'report_id': str(report_id),
                'status': 'completed',
                'file_path': file_path,
                'file_size': file_size
            }
            
        except Exception as e:
            logger.error(f"Report generation failed: {e}", exc_info=True)
            self._update_report_status(report_id, 'failed', error=str(e))
            return {
                'report_id': str(report_id),
                'status': 'failed',
                'error': str(e)
            }
    
    def _create_report_record(
        self, user_id: UUID, report_type: str, parameters: Dict
    ) -> UUID:
        """Create report record in database"""
        cur = self.db.cursor()
        
        # Get user's organization
        cur.execute("SELECT organization_id FROM users WHERE id = %s", (str(user_id),))
        user = cur.fetchone()
        org_id = user['organization_id'] if user else None
        
        from uuid import uuid4
        report_id = uuid4()
        
        cur.execute("""
            INSERT INTO reports (
                id, user_id, organization_id, report_type, report_name,
                date_range_start, date_range_end, watchlist_item_id, status
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pending')
            RETURNING id
        """, (
            str(report_id), str(user_id), str(org_id) if org_id else None,
            report_type, f"{report_type}_{datetime.utcnow().strftime('%Y%m%d')}",
            parameters.get('date_start'), parameters.get('date_end'),
            str(parameters.get('watchlist_id')) if parameters.get('watchlist_id') else None
        ))
        
        self.db.commit()
        return report_id
    
    def _update_report_status(
        self, report_id: UUID, status: str,
        file_path: str = None, file_size: int = None, error: str = None
    ):
        """Update report status"""
        cur = self.db.cursor()
        
        if status == 'completed':
            cur.execute("""
                UPDATE reports SET 
                    status = %s, file_path = %s, file_size_bytes = %s,
                    generated_at = NOW(), expires_at = NOW() + INTERVAL '30 days'
                WHERE id = %s
            """, (status, file_path, file_size, str(report_id)))
        elif status == 'failed':
            cur.execute("""
                UPDATE reports SET status = %s, error_message = %s
                WHERE id = %s
            """, (status, error, str(report_id)))
        else:
            cur.execute("UPDATE reports SET status = %s WHERE id = %s", (status, str(report_id)))
        
        self.db.commit()
    
    def _generate_weekly_digest(
        self, user_id: UUID, report_id: UUID, parameters: Dict
    ) -> str:
        """Generate weekly alert digest report"""
        cur = self.db.cursor()
        
        # Get alerts from last 7 days
        cur.execute("""
            SELECT
                a.*,
                w.brand_name as watched_brand,
                w.nice_class_numbers as watched_classes
            FROM alerts_mt a
            JOIN watchlist_mt w ON a.watchlist_item_id = w.id
            WHERE a.user_id = %s
            AND a.created_at > NOW() - INTERVAL '7 days'
            ORDER BY a.severity DESC, a.created_at DESC
        """, (str(user_id),))
        
        alerts = [dict(row) for row in cur.fetchall()]
        
        # Get user info
        cur.execute("SELECT first_name, last_name, email FROM users WHERE id = %s", (str(user_id),))
        user = dict(cur.fetchone())
        
        # Generate PDF
        return self._create_digest_pdf(user, alerts, 'weekly', report_id)
    
    def _generate_monthly_summary(
        self, user_id: UUID, report_id: UUID, parameters: Dict
    ) -> str:
        """Generate monthly summary report"""
        cur = self.db.cursor()
        
        # Get alerts from last 30 days with statistics
        cur.execute("""
            SELECT
                a.*,
                w.brand_name as watched_brand,
                w.nice_class_numbers as watched_classes
            FROM alerts_mt a
            JOIN watchlist_mt w ON a.watchlist_item_id = w.id
            WHERE a.user_id = %s
            AND a.created_at > NOW() - INTERVAL '30 days'
            ORDER BY a.severity DESC, a.created_at DESC
        """, (str(user_id),))
        
        alerts = [dict(row) for row in cur.fetchall()]
        
        # Get statistics
        cur.execute("""
            SELECT
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE severity = 'critical') as critical,
                COUNT(*) FILTER (WHERE severity = 'very_high') as very_high,
                COUNT(*) FILTER (WHERE severity = 'high') as high,
                COUNT(*) FILTER (WHERE severity = 'medium') as medium,
                COUNT(*) FILTER (WHERE severity = 'low') as low,
                COUNT(*) FILTER (WHERE status = 'resolved') as resolved
            FROM alerts_mt
            WHERE user_id = %s
            AND created_at > NOW() - INTERVAL '30 days'
        """, (str(user_id),))
        
        stats = dict(cur.fetchone())
        
        # Get user info
        cur.execute("SELECT first_name, last_name, email FROM users WHERE id = %s", (str(user_id),))
        user = dict(cur.fetchone())
        
        return self._create_summary_pdf(user, alerts, stats, 'monthly', report_id)
    
    def _generate_watchlist_status(
        self, user_id: UUID, report_id: UUID, parameters: Dict
    ) -> str:
        """Generate watchlist status report"""
        cur = self.db.cursor()
        
        # Get all watchlist items with stats
        cur.execute("""
            SELECT
                w.*,
                COUNT(a.id) as alert_count,
                COUNT(a.id) FILTER (WHERE a.status = 'new') as new_alerts,
                MAX(a.overall_risk_score) as highest_risk,
                MAX(a.created_at) as latest_alert
            FROM watchlist_mt w
            LEFT JOIN alerts_mt a ON w.id = a.watchlist_item_id
            WHERE w.user_id = %s AND w.is_active = TRUE
            GROUP BY w.id
            ORDER BY highest_risk DESC NULLS LAST, w.brand_name
        """, (str(user_id),))
        
        items = [dict(row) for row in cur.fetchall()]
        
        # Get user info
        cur.execute("SELECT first_name, last_name, email FROM users WHERE id = %s", (str(user_id),))
        user = dict(cur.fetchone())
        
        return self._create_watchlist_pdf(user, items, report_id)
    
    def _generate_single_trademark_report(
        self, user_id: UUID, report_id: UUID, watchlist_id: UUID, parameters: Dict
    ) -> str:
        """Generate detailed report for single trademark"""
        cur = self.db.cursor()
        
        # Get watchlist item
        cur.execute("""
            SELECT * FROM watchlist_mt WHERE id = %s AND user_id = %s
        """, (str(watchlist_id), str(user_id)))
        
        item = cur.fetchone()
        if not item:
            raise ValueError("Watchlist item not found")
        item = dict(item)
        
        # Get all alerts for this item
        cur.execute("""
            SELECT * FROM alerts_mt
            WHERE watchlist_item_id = %s
            ORDER BY created_at DESC
        """, (str(watchlist_id),))
        
        alerts = [dict(row) for row in cur.fetchall()]
        
        # Get user info
        cur.execute("SELECT first_name, last_name, email FROM users WHERE id = %s", (str(user_id),))
        user = dict(cur.fetchone())
        
        return self._create_trademark_pdf(user, item, alerts, report_id)
    
    def _generate_portfolio_report(
        self, user_id: UUID, report_id: UUID, parameters: Dict
    ) -> str:
        """Generate full portfolio report (all trademarks + all alerts)"""
        cur = self.db.cursor()
        
        # Get all watchlist items
        cur.execute("""
            SELECT * FROM watchlist_mt
            WHERE user_id = %s AND is_active = TRUE
            ORDER BY brand_name
        """, (str(user_id),))
        
        items = [dict(row) for row in cur.fetchall()]
        
        # Get all alerts
        cur.execute("""
            SELECT
                a.*,
                w.brand_name as watched_brand
            FROM alerts_mt a
            JOIN watchlist_mt w ON a.watchlist_item_id = w.id
            WHERE a.user_id = %s
            ORDER BY a.created_at DESC
        """, (str(user_id),))
        
        alerts = [dict(row) for row in cur.fetchall()]
        
        # Get user info
        cur.execute("SELECT first_name, last_name, email FROM users WHERE id = %s", (str(user_id),))
        user = dict(cur.fetchone())
        
        return self._create_portfolio_pdf(user, items, alerts, report_id)
    
    # ==========================================
    # PDF Generation Methods
    # ==========================================
    
    def _create_digest_pdf(
        self, user: Dict, alerts: List[Dict], period: str, report_id: UUID
    ) -> str:
        """Create PDF digest report"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            # Fallback to simple text file if reportlab not available
            return self._create_text_report(user, alerts, period, report_id)
        
        filename = f"digest_{period}_{report_id}.pdf"
        filepath = self.output_dir / filename
        
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'Title', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=30
        )
        story.append(Paragraph(
            f"Trademark Alert {period.title()} Digest",
            title_style
        ))
        
        # User info
        story.append(Paragraph(
            f"Prepared for: {user.get('first_name', '')} {user.get('last_name', '')}",
            styles['Normal']
        ))
        story.append(Paragraph(
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            styles['Normal']
        ))
        story.append(Spacer(1, 20))
        
        # Summary
        critical = sum(1 for a in alerts if a.get('severity') == 'critical')
        high = sum(1 for a in alerts if a.get('severity') == 'high')
        
        story.append(Paragraph(f"<b>Summary:</b>", styles['Heading2']))
        story.append(Paragraph(f"Total Alerts: {len(alerts)}", styles['Normal']))
        story.append(Paragraph(f"Critical: {critical}", styles['Normal']))
        story.append(Paragraph(f"High: {high}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Alerts table
        if alerts:
            story.append(Paragraph("<b>Alert Details:</b>", styles['Heading2']))
            
            table_data = [['Brand', 'Conflict', 'Score', 'Severity', 'Date']]
            
            for alert in alerts[:50]:  # Limit to 50
                table_data.append([
                    alert.get('watched_brand', '')[:20],
                    alert.get('conflicting_name', '')[:20],
                    f"{alert.get('overall_risk_score', 0):.0%}",
                    alert.get('severity', '').upper(),
                    alert.get('created_at', datetime.utcnow()).strftime('%Y-%m-%d') if alert.get('created_at') else ''
                ])
            
            table = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 0.8*inch, 0.8*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("No alerts during this period.", styles['Normal']))
        
        doc.build(story)
        return str(filepath)
    
    def _create_text_report(
        self, user: Dict, alerts: List[Dict], period: str, report_id: UUID
    ) -> str:
        """Fallback: Create simple text report"""
        filename = f"digest_{period}_{report_id}.txt"
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"TRADEMARK ALERT {period.upper()} DIGEST\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Prepared for: {user.get('first_name', '')} {user.get('last_name', '')}\n")
            f.write(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n\n")
            
            f.write(f"Total Alerts: {len(alerts)}\n")
            f.write(f"Critical: {sum(1 for a in alerts if a.get('severity') == 'critical')}\n")
            f.write(f"High: {sum(1 for a in alerts if a.get('severity') == 'high')}\n\n")
            
            f.write("-" * 50 + "\n")
            f.write("ALERT DETAILS\n")
            f.write("-" * 50 + "\n\n")
            
            for alert in alerts:
                f.write(f"Brand: {alert.get('watched_brand', 'N/A')}\n")
                f.write(f"Conflict: {alert.get('conflicting_name', 'N/A')}\n")
                f.write(f"Score: {alert.get('overall_risk_score', 0):.0%}\n")
                f.write(f"Severity: {alert.get('severity', 'N/A').upper()}\n")
                f.write(f"Date: {alert.get('created_at', '')}\n")
                f.write("\n")
        
        return str(filepath)
    
    def _create_summary_pdf(self, user, alerts, stats, period, report_id):
        """Create summary PDF - similar to digest with more stats"""
        return self._create_digest_pdf(user, alerts, period, report_id)
    
    def _create_watchlist_pdf(self, user, items, report_id):
        """Create watchlist status PDF"""
        return self._create_text_report(
            user, 
            [{'watched_brand': i.get('brand_name'), 'severity': 'info', 
              'conflicting_name': f"{i.get('alert_count', 0)} alerts",
              'overall_risk_score': i.get('highest_risk') or 0,
              'created_at': i.get('created_at')} for i in items],
            'watchlist_status',
            report_id
        )
    
    def _create_trademark_pdf(self, user, item, alerts, report_id):
        """Create single trademark report PDF"""
        return self._create_digest_pdf(
            user, alerts, f"trademark_{item.get('brand_name', 'unknown')}", report_id
        )
    
    def _create_portfolio_pdf(self, user, items, alerts, report_id):
        """Create full portfolio PDF"""
        return self._create_digest_pdf(user, alerts, 'portfolio', report_id)
    
    # ==========================================
    # Excel Report Methods
    # ==========================================
    
    def generate_excel_report(
        self, user_id: UUID, report_type: str, parameters: Dict = None
    ) -> str:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Fill, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl required for Excel reports: pip install openpyxl")
        
        parameters = parameters or {}
        report_id = self._create_report_record(user_id, report_type, parameters)
        
        try:
            self._update_report_status(report_id, 'generating')
            
            cur = self.db.cursor()
            
            # Get alerts
            cur.execute("""
                SELECT
                    a.*,
                    w.brand_name as watched_brand,
                    w.nice_class_numbers as watched_classes
                FROM alerts_mt a
                JOIN watchlist_mt w ON a.watchlist_item_id = w.id
                WHERE a.user_id = %s
                ORDER BY a.created_at DESC
            """, (str(user_id),))
            
            alerts = [dict(row) for row in cur.fetchall()]
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Alerts"
            
            # Header
            headers = ['Brand', 'Conflict', 'App No', 'Status', 'Score', 'Severity', 'Date']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            
            # Data
            for row, alert in enumerate(alerts, 2):
                ws.cell(row=row, column=1, value=alert.get('watched_brand', ''))
                ws.cell(row=row, column=2, value=alert.get('conflicting_name', ''))
                ws.cell(row=row, column=3, value=alert.get('conflicting_application_no', ''))
                ws.cell(row=row, column=4, value=alert.get('conflicting_status', ''))
                ws.cell(row=row, column=5, value=f"{alert.get('overall_risk_score', 0):.0%}")
                ws.cell(row=row, column=6, value=alert.get('severity', '').upper())
                ws.cell(row=row, column=7, value=str(alert.get('created_at', ''))[:10])
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            
            # Save
            filename = f"alerts_{report_id}.xlsx"
            filepath = self.output_dir / filename
            wb.save(str(filepath))
            
            file_size = os.path.getsize(str(filepath))
            self._update_report_status(report_id, 'completed', str(filepath), file_size)
            
            return str(filepath)
            
        except Exception as e:
            self._update_report_status(report_id, 'failed', error=str(e))
            raise


# ==========================================
# CLI Entry Point
# ==========================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate reports")
    parser.add_argument("--user-id", type=str, required=True, help="User ID")
    parser.add_argument("--type", type=str, required=True,
                        choices=['weekly_digest', 'monthly_summary', 'watchlist_status', 'full_portfolio'],
                        help="Report type")
    parser.add_argument("--format", type=str, default='pdf', choices=['pdf', 'xlsx'],
                        help="Output format")
    
    args = parser.parse_args()
    
    logging.basicConfig(level=logging.INFO)
    
    generator = ReportGenerator()
    
    if args.format == 'xlsx':
        result = generator.generate_excel_report(UUID(args.user_id), args.type)
    else:
        result = generator.generate_report(UUID(args.user_id), args.type)
    
    print(f"Report generated: {result}")
