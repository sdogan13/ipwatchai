"""Watchlist service helpers used by HTTP route modules."""

import io
import json
import os
import re
from datetime import date as date_type
from uuid import UUID, uuid4

import pandas as pd
import psycopg2
from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from psycopg2 import sql as psql

from config.settings import PROJECT_ROOT, settings
from database.crud import Database, WatchlistCRUD
from models.schemas import (
    FileUploadErrorItem,
    FileUploadResult,
    FileUploadSkippedItem,
    FileUploadSummary,
    FileUploadWarning,
    WatchlistItemCreate,
    WatchlistItemResponse,
)

WATCHLIST_LOGOS_DIR = os.path.join(settings.paths.upload_dir, "watchlist_logos")


def _normalize_ratio_threshold(value) -> float:
    """Accept either decimal (0.8) or percentage (80) threshold input."""
    return value / 100.0 if value > 1.0 else float(value)


def _days_until(target_date, today_factory=None):
    """Return the number of days from today until the target date."""
    if target_date is None:
        return None
    if today_factory is None:
        today_factory = date_type.today
    return (target_date - today_factory()).days


def _load_conflict_summaries(
    db,
    item_ids,
    threshold,
    logger=None,
    today_factory=None,
    appeals_only: bool = False,
    status_filter=None,
):
    """Load per-item conflict summary data for watchlist list responses."""
    if not item_ids:
        return {}

    try:
        cur = db.cursor()
        resolved_or_dismissed_appeals = appeals_only and status_filter in ("resolved", "dismissed")
        params = (item_ids,)

        if resolved_or_dismissed_appeals:
            where_clause = """
                a.watchlist_item_id = ANY(%s::uuid[])
                AND a.status = %s
            """
            params += (status_filter,)
            nearest_deadline_expr = "MIN(t.appeal_deadline) as nearest_deadline,"
            severity_rank_expr = """
                MAX(CASE a.severity
                    WHEN 'critical'  THEN 5
                    WHEN 'very_high' THEN 4
                    WHEN 'high'      THEN 3
                    WHEN 'medium'    THEN 2
                    WHEN 'low'       THEN 1
                    ELSE 0
                END) AS highest_severity_rank,
            """
        else:
            where_clause = """
                a.watchlist_item_id = ANY(%s::uuid[])
                AND a.status NOT IN ('dismissed', 'resolved')
                AND (t.appeal_deadline IS NULL OR t.appeal_deadline >= CURRENT_DATE)
                AND a.overall_risk_score >= %s
            """
            params += (threshold,)
            nearest_deadline_expr = (
                "MIN(t.appeal_deadline) FILTER (WHERE t.appeal_deadline > CURRENT_DATE) as nearest_deadline,"
            )
            severity_rank_expr = """
                MAX(CASE a.severity
                    WHEN 'critical'  THEN 5
                    WHEN 'very_high' THEN 4
                    WHEN 'high'      THEN 3
                    WHEN 'medium'    THEN 2
                    WHEN 'low'       THEN 1
                    ELSE 0
                END) FILTER (WHERE a.status NOT IN ('dismissed', 'resolved')) AS highest_severity_rank,
            """

        cur.execute(
            f"""
            SELECT
                a.watchlist_item_id,
                COUNT(*) as total_conflicts,
                COUNT(*) FILTER (WHERE t.final_status = 'Başvuruldu' AND t.bulletin_date IS NULL) as pre_publication_count,
                COUNT(*) FILTER (WHERE t.appeal_deadline > CURRENT_DATE AND t.appeal_deadline <= CURRENT_DATE + INTERVAL '7 days') as critical_count,
                COUNT(*) FILTER (WHERE t.appeal_deadline > CURRENT_DATE + INTERVAL '7 days' AND t.appeal_deadline <= CURRENT_DATE + INTERVAL '30 days') as urgent_count,
                COUNT(*) FILTER (WHERE t.appeal_deadline > CURRENT_DATE + INTERVAL '30 days') as active_count,
                {nearest_deadline_expr}
                {severity_rank_expr}
                COUNT(*) FILTER (WHERE a.severity = 'critical') as sev_critical,
                COUNT(*) FILTER (WHERE a.severity = 'very_high') as sev_very_high,
                COUNT(*) FILTER (WHERE a.severity = 'high') as sev_high,
                COUNT(*) FILTER (WHERE a.severity = 'medium') as sev_medium,
                COUNT(*) FILTER (WHERE a.severity = 'low') as sev_low
            FROM alerts_mt a
            LEFT JOIN trademarks t ON a.conflicting_trademark_id = t.id
            WHERE {where_clause}
            GROUP BY a.watchlist_item_id
        """,
            params,
        )
        severity_map = {
            5: "critical",
            4: "very_high",
            3: "high",
            2: "medium",
            1: "low",
        }
        summaries = {}
        for row in cur.fetchall():
            nearest = row["nearest_deadline"]
            summaries[str(row["watchlist_item_id"])] = {
                "total": row["total_conflicts"],
                "pre_publication": row["pre_publication_count"],
                "active_critical": row["critical_count"],
                "active_urgent": row["urgent_count"],
                "active": row["active_count"],
                "nearest_deadline": nearest.isoformat() if nearest else None,
                "nearest_deadline_days": _days_until(nearest, today_factory=today_factory),
                "highest_severity": severity_map.get(row["highest_severity_rank"]),
                "sev_critical": row["sev_critical"],
                "sev_very_high": row["sev_very_high"],
                "sev_high": row["sev_high"],
                "sev_medium": row["sev_medium"],
                "sev_low": row["sev_low"],
            }
        return summaries
    except Exception:
        if logger:
            logger.exception(
                "Failed to load conflict summaries; watchlist will render without badge counts"
            )
        return {}


async def get_watchlist_stats_summary(
    current_user,
    min_score: float = 0.0,
    database_factory=None,
    today_factory=None,
):
    """Return aggregate stats for the current organization's watchlist."""
    if database_factory is None:
        database_factory = Database

    norm_score = _normalize_ratio_threshold(min_score)
    with database_factory() as db:
        cur = db.cursor()
        cur.execute(
            f"""
            SELECT
                COUNT(DISTINCT w.id) AS total_items,
                COUNT(DISTINCT w.id) FILTER (WHERE w.is_active = TRUE) AS active_items,
                COUNT(DISTINCT w.id) FILTER (WHERE a.id IS NOT NULL
                    AND a.status NOT IN ('dismissed', 'resolved')
                    AND a.overall_risk_score >= {norm_score}
                    AND (t.appeal_deadline IS NOT NULL AND t.appeal_deadline >= CURRENT_DATE)) AS items_with_threats,
                COUNT(a.id) FILTER (WHERE a.severity = 'critical' AND a.status NOT IN ('dismissed', 'resolved')
                    AND a.overall_risk_score >= {norm_score}
                    AND (t.appeal_deadline IS NOT NULL AND t.appeal_deadline >= CURRENT_DATE)) AS critical_threats,
                COUNT(a.id) FILTER (WHERE a.severity = 'high' AND a.status NOT IN ('dismissed', 'resolved')
                    AND a.overall_risk_score >= {norm_score}
                    AND (t.appeal_deadline IS NOT NULL AND t.appeal_deadline >= CURRENT_DATE)) AS high_threats,
                COUNT(a.id) FILTER (WHERE a.severity = 'medium' AND a.status NOT IN ('dismissed', 'resolved')
                    AND a.overall_risk_score >= {norm_score}
                    AND (t.appeal_deadline IS NOT NULL AND t.appeal_deadline >= CURRENT_DATE)) AS medium_threats,
                COUNT(a.id) FILTER (WHERE a.severity = 'low' AND a.status NOT IN ('dismissed', 'resolved')
                    AND a.overall_risk_score >= {norm_score}
                    AND (t.appeal_deadline IS NOT NULL AND t.appeal_deadline >= CURRENT_DATE)) AS low_threats,
                COUNT(a.id) FILTER (WHERE a.status = 'new'
                    AND a.overall_risk_score >= {norm_score}
                    AND (t.appeal_deadline IS NOT NULL AND t.appeal_deadline >= CURRENT_DATE)) AS new_alerts,
                MIN(t.appeal_deadline) FILTER (WHERE t.appeal_deadline > CURRENT_DATE
                    AND a.overall_risk_score >= {norm_score}
                    AND a.status NOT IN ('dismissed', 'resolved')) AS nearest_deadline,
                COUNT(DISTINCT w.id) FILTER (
                    WHERE (my_tm.application_date IS NOT NULL
                           AND my_tm.application_date + INTERVAL '10 years 6 months' <= CURRENT_DATE + INTERVAL '12 months')
                       OR (my_tm.application_date IS NULL
                           AND w.customer_application_no IS NOT NULL
                           AND left(w.customer_application_no, 4) ~ '^[0-9]{4}$'
                           AND EXTRACT(YEAR FROM CURRENT_DATE)::int - left(w.customer_application_no, 4)::int >= 9)
                ) AS renewal_count
            FROM watchlist_mt w
            LEFT JOIN alerts_mt a ON w.id = a.watchlist_item_id
            LEFT JOIN trademarks t ON a.conflicting_trademark_id = t.id
            LEFT JOIN trademarks my_tm ON w.customer_application_no = my_tm.application_no
            WHERE w.organization_id = %s AND w.is_active = TRUE
        """,
            (str(current_user.organization_id),),
        )
        row = cur.fetchone()

    nearest = row["nearest_deadline"]
    return {
        "total_items": row["total_items"],
        "active_items": row["active_items"],
        "items_with_threats": row["items_with_threats"],
        "critical_threats": row["critical_threats"],
        "high_threats": row["high_threats"],
        "medium_threats": row["medium_threats"],
        "low_threats": row["low_threats"],
        "new_alerts": row["new_alerts"],
        "nearest_deadline": nearest.isoformat() if nearest else None,
        "nearest_deadline_days": _days_until(nearest, today_factory=today_factory),
        "renewal_count": row["renewal_count"],
    }


async def get_watchlist_page(
    current_user,
    page: int = 1,
    page_size: int = 20,
    active_only: bool = True,
    search=None,
    sort=None,
    renewal_only: bool = False,
    appeals_only: bool = False,
    status_filter=None,
    threshold: float = 0.50,
    tm_status=None,
    database_factory=None,
    watchlist_crud=None,
    logger=None,
    today_factory=None,
):
    """Return the paginated watchlist page for the current organization."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    norm_threshold = _normalize_ratio_threshold(threshold)

    with database_factory() as db:
        items, total = watchlist_crud.get_by_organization(
            db,
            current_user.organization_id,
            active_only,
            page,
            page_size,
            search=search,
            sort_by=sort,
            renewal_only=renewal_only,
            appeals_only=appeals_only,
            status_filter=status_filter,
            threshold=norm_threshold,
            tm_status=tm_status,
        )

        conflict_summaries = _load_conflict_summaries(
            db,
            [item["id"] for item in items],
            norm_threshold,
            logger=logger,
            today_factory=today_factory,
            appeals_only=appeals_only,
            status_filter=status_filter,
        )

    response_items = []
    for item in items:
        item_payload = dict(item)
        item_payload["conflict_summary"] = conflict_summaries.get(str(item["id"]))
        response_items.append(WatchlistItemResponse(**item_payload).model_dump())

    return {
        "items": response_items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }


def _parse_embedding_vector(value):
    """Convert stored vector text back into a float list."""
    if not value:
        return None
    if isinstance(value, list):
        return value
    value = value.strip()
    if value.startswith("[") and value.endswith("]"):
        return [float(item) for item in value[1:-1].split(",") if item.strip()]
    return None


def _raise_watchlist_item_not_found():
    """Raise the standard not-found error for watchlist items."""
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="Item not found",
    )


def _ensure_logo_tracking_allowed(
    db,
    user_id,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Ensure the user plan allows logo tracking for manual watchlist entries."""
    if user_plan_getter is None:
        from utils.subscription import get_user_plan

        user_plan_getter = get_user_plan
    if plan_limit_getter is None:
        from utils.subscription import get_plan_limit

        plan_limit_getter = get_plan_limit

    plan = user_plan_getter(db, str(user_id))
    can_track = plan_limit_getter(plan["plan_name"], "can_track_logos")
    if not can_track:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "error": "upgrade_required",
                "message": "Logo tracking requires a paid plan.",
            },
        )


def _raise_watchlist_limit_exceeded(max_items, current_count, current_plan=None):
    """Raise the standard watchlist capacity error."""
    detail = {
        "error": "limit_exceeded",
        "message": f"Izleme listesi limitinize ulastiniz ({max_items}). Daha fazla eklemek icin planinizi yukseltin.",
        "current_count": current_count,
        "max_items": max_items,
    }
    if current_plan:
        detail["current_plan"] = current_plan
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail=detail,
    )


def _get_watchlist_capacity(
    db,
    current_user,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Return plan-derived watchlist capacity numbers for the current user."""
    if user_plan_getter is None:
        from utils.subscription import get_user_plan

        user_plan_getter = get_user_plan
    if plan_limit_getter is None:
        from utils.subscription import get_plan_limit

        plan_limit_getter = get_plan_limit

    plan_info = user_plan_getter(db, str(current_user.id))
    plan_name = plan_info.get("plan_name", "free")
    max_items = plan_limit_getter(plan_name, "max_watchlist_items")

    cur = db.cursor()
    cur.execute(
        "SELECT COUNT(*) FROM watchlist_mt WHERE organization_id = %s AND is_active = TRUE",
        (str(current_user.organization_id),),
    )
    current_count = cur.fetchone()["count"]
    remaining_slots = max(0, max_items - current_count)

    return {
        "plan_name": plan_name,
        "max_items": max_items,
        "current_count": current_count,
        "remaining_slots": remaining_slots,
    }


def _load_existing_watchlist_application_numbers(db, organization_id):
    """Return active watchlist application numbers already tracked for an org."""
    cur = db.cursor()
    cur.execute(
        "SELECT customer_application_no FROM watchlist_mt WHERE organization_id = %s AND customer_application_no IS NOT NULL",
        (str(organization_id),),
    )
    return {
        str(row["customer_application_no"]).strip() for row in cur.fetchall()
    }


def _require_portfolio_lookup_params(data):
    """Require at least one holder or attorney identifier for portfolio endpoints."""
    if not getattr(data, "holder_id", None) and not getattr(data, "attorney_no", None):
        raise HTTPException(status_code=400, detail="holder_id or attorney_no required")


def _find_column_match(columns, variants):
    """Return the first matching normalized column name from the provided variants."""
    for variant in variants:
        if variant in columns:
            return variant
    return None


def _ensure_portfolio_access_allowed(
    db,
    user_id,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Ensure the user's plan can import from holder or attorney portfolios."""
    if user_plan_getter is None:
        from utils.subscription import get_user_plan

        user_plan_getter = get_user_plan
    if plan_limit_getter is None:
        from utils.subscription import get_plan_limit

        plan_limit_getter = get_plan_limit

    plan = user_plan_getter(db, str(user_id))
    can_view = plan_limit_getter(plan["plan_name"], "can_view_holder_portfolio")
    if not can_view:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "error": "upgrade_required",
                "message": "Portfoy erisimi icin Business veya ustu plan gereklidir.",
                "current_plan": plan["plan_name"],
            },
        )


def _get_auto_scan_limit(
    db,
    user_id,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Return the plan-based auto scan limit or raise the standard upgrade error."""
    if user_plan_getter is None:
        from utils.subscription import get_user_plan

        user_plan_getter = get_user_plan
    if plan_limit_getter is None:
        from utils.subscription import get_plan_limit

        plan_limit_getter = get_plan_limit

    plan = user_plan_getter(db, str(user_id))
    scan_max = plan_limit_getter(plan["plan_name"], "auto_scan_max_items")
    if scan_max == 0:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "error": "upgrade_required",
                "message": "Otomatik tarama icin planinizi yukseltin.",
                "current_plan": plan["plan_name"],
            },
        )
    return scan_max, plan


async def create_watchlist_item_record(
    data,
    current_user,
    database_factory=None,
    watchlist_crud=None,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Create a watchlist item and return the created record plus scan target."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    raw_app_no = getattr(data, "application_no", None)
    app_no = str(raw_app_no).strip() if raw_app_no else None

    if app_no:
        with database_factory() as db_dup:
            cur = db_dup.cursor()
            cur.execute(
                "SELECT 1 FROM watchlist_mt WHERE organization_id = %s AND customer_application_no = %s AND is_active = TRUE",
                (str(current_user.organization_id), app_no),
            )
            if cur.fetchone():
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Izleme listesinde zaten var",
                )

    if getattr(data, "monitor_visual", False) and not app_no:
        with database_factory() as db_check:
            _ensure_logo_tracking_allowed(
                db_check,
                current_user.id,
                user_plan_getter=user_plan_getter,
                plan_limit_getter=plan_limit_getter,
            )

    with database_factory() as db:
        try:
            tm_ai = None

            if app_no:
                cur = db.cursor()
                cur.execute(
                    """
                    SELECT image_path,
                           image_embedding::text, dinov2_embedding::text,
                           color_histogram::text, logo_ocr_text, text_embedding::text
                    FROM trademarks
                    WHERE application_no = %s
                    LIMIT 1
                """,
                    (app_no,),
                )
                tm_ai = cur.fetchone()

            if tm_ai:
                item = watchlist_crud.create_with_embeddings(
                    db,
                    current_user.organization_id,
                    current_user.id,
                    data,
                    logo_path=tm_ai.get("image_path") or None,
                    logo_embedding=_parse_embedding_vector(tm_ai.get("image_embedding")),
                    logo_dinov2_embedding=_parse_embedding_vector(
                        tm_ai.get("dinov2_embedding")
                    ),
                    logo_color_histogram=_parse_embedding_vector(
                        tm_ai.get("color_histogram")
                    ),
                    logo_ocr_text=tm_ai.get("logo_ocr_text"),
                    text_embedding=_parse_embedding_vector(tm_ai.get("text_embedding")),
                )
            else:
                item = watchlist_crud.create(
                    db,
                    current_user.organization_id,
                    current_user.id,
                    data,
                )

            return {
                "item": item,
                "scan_item_id": UUID(str(item["id"])),
            }
        except ValueError as exc:
            if str(exc) == "Organization has reached maximum watchlist items limit":
                capacity = _get_watchlist_capacity(
                    db,
                    current_user,
                    user_plan_getter=user_plan_getter,
                    plan_limit_getter=plan_limit_getter,
                )
                _raise_watchlist_limit_exceeded(
                    capacity["max_items"],
                    capacity["current_count"],
                    current_plan=capacity["plan_name"],
                )
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(exc),
            ) from exc
        except psycopg2.errors.UniqueViolation as exc:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Izleme listesinde zaten var",
            ) from exc


async def get_watchlist_item_detail(
    item_id,
    current_user,
    database_factory=None,
    watchlist_crud=None,
):
    """Load a single watchlist item or raise a 404."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    with database_factory() as db:
        item = watchlist_crud.get_by_id(db, item_id, current_user.organization_id)
        if not item:
            _raise_watchlist_item_not_found()
        return item


async def update_watchlist_item_record(
    item_id,
    data,
    current_user,
    database_factory=None,
    watchlist_crud=None,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Update a watchlist item and return the refreshed record."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    if getattr(data, "monitor_visual", None) is True:
        with database_factory() as db_check:
            _ensure_logo_tracking_allowed(
                db_check,
                current_user.id,
                user_plan_getter=user_plan_getter,
                plan_limit_getter=plan_limit_getter,
            )

    with database_factory() as db:
        item = watchlist_crud.update(db, item_id, current_user.organization_id, data)
        if not item:
            _raise_watchlist_item_not_found()
        return item


async def delete_watchlist_item_record(
    item_id,
    current_user,
    database_factory=None,
    watchlist_crud=None,
):
    """Delete a watchlist item and report how many alerts were removed."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    with database_factory() as db:
        cur = db.cursor()
        cur.execute(
            """
            DELETE FROM alerts_mt
            WHERE watchlist_item_id = %s
        """,
            (str(item_id),),
        )
        deleted_alerts = cur.rowcount

        success = watchlist_crud.delete(db, item_id, current_user.organization_id)
        if not success:
            _raise_watchlist_item_not_found()

        db.commit()
        return {
            "success": True,
            "message": f"Marka ve {deleted_alerts} uyari silindi",
        }


async def prepare_watchlist_item_scan(
    item_id,
    current_user,
    database_factory=None,
    watchlist_crud=None,
):
    """Validate a watchlist item exists before enqueuing a scan."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    with database_factory() as db:
        item = watchlist_crud.get_by_id(db, item_id, current_user.organization_id)
        if not item:
            _raise_watchlist_item_not_found()

    return {
        "success": True,
        "message": "Scan triggered",
        "item_id": item_id,
    }


async def prepare_watchlist_scan_all(
    current_user,
    database_factory=None,
    watchlist_crud=None,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Prepare an organization-wide scan and return the queued item ids."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    with database_factory() as db:
        scan_max, _ = _get_auto_scan_limit(
            db,
            current_user.id,
            user_plan_getter=user_plan_getter,
            plan_limit_getter=plan_limit_getter,
        )
        _, total = watchlist_crud.get_by_organization(
            db, current_user.organization_id, active_only=True, page_size=1
        )
        items, _ = watchlist_crud.get_by_organization(
            db, current_user.organization_id, active_only=True, page_size=max(total, 1)
        )

    if not items:
        return {
            "success": True,
            "message": "Izleme listesinde taranacak marka yok",
            "item_ids": [],
        }

    items_to_scan = items[:scan_max] if scan_max < 999999 else items
    message = f"{len(items_to_scan)} marka taramaya alindi (toplam: {total})"
    if len(items_to_scan) < len(items):
        message += f" Ã¢â‚¬â€ plan limitiniz nedeniyle {scan_max} marka tarandi"

    return {
        "success": True,
        "message": message,
        "item_ids": [UUID(str(item["id"])) for item in items_to_scan],
    }


async def get_watchlist_scan_status(current_user=None, next_scan_time_getter=None):
    """Return the watchlist auto-scan schedule metadata."""
    if next_scan_time_getter is None:
        from workers.scheduler import get_next_scan_time

        next_scan_time_getter = get_next_scan_time

    return {
        "auto_scan_enabled": True,
        "schedule": "Daily at 03:00",
        "next_scan_at": next_scan_time_getter(),
    }


async def delete_all_watchlist_records(current_user, database_factory=None):
    """Delete all watchlist items and alerts for an organization."""
    if database_factory is None:
        database_factory = Database

    with database_factory() as db:
        cur = db.cursor()
        org_id = str(current_user.organization_id)

        cur.execute(
            """
            DELETE FROM alerts_mt WHERE organization_id = %s
        """,
            (org_id,),
        )
        deleted_alerts = cur.rowcount

        cur.execute(
            """
            DELETE FROM watchlist_mt WHERE organization_id = %s
        """,
            (org_id,),
        )
        deleted_items = cur.rowcount

        db.commit()

    return {
        "success": True,
        "message": f"{deleted_items} marka ve {deleted_alerts} uyari silindi",
    }


async def prepare_watchlist_rescan(
    current_user,
    database_factory=None,
    watchlist_crud=None,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Clear existing alerts and prepare a fresh organization-wide rescan."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    with database_factory() as db:
        rescan_max, _ = _get_auto_scan_limit(
            db,
            current_user.id,
            user_plan_getter=user_plan_getter,
            plan_limit_getter=plan_limit_getter,
        )

        cur = db.cursor()
        org_id = str(current_user.organization_id)

        cur.execute(
            """
            DELETE FROM alerts_mt WHERE organization_id = %s
        """,
            (org_id,),
        )
        cleared_alerts = cur.rowcount

        cur.execute(
            """
            UPDATE watchlist_mt SET last_scan_at = NULL WHERE organization_id = %s
        """,
            (org_id,),
        )

        _, total = watchlist_crud.get_by_organization(
            db, current_user.organization_id, active_only=True, page_size=1
        )
        items, _ = watchlist_crud.get_by_organization(
            db, current_user.organization_id, active_only=True, page_size=max(total, 1)
        )

        db.commit()

    if not items:
        return {
            "success": True,
            "message": f"Eski {cleared_alerts} uyari silindi. Taranacak marka yok.",
            "item_ids": [],
        }

    items_to_scan = items[:rescan_max] if rescan_max < 999999 else items
    return {
        "success": True,
        "message": f"Eski {cleared_alerts} uyari silindi. {len(items_to_scan)} marka yeniden taramaya alindi.",
        "item_ids": [UUID(str(item["id"])) for item in items_to_scan],
    }


async def update_watchlist_bulk_thresholds(
    threshold,
    current_user,
    database_factory=None,
):
    """Update the alert threshold for all active watchlist items in the organization."""
    if database_factory is None:
        database_factory = Database

    with database_factory() as db:
        cur = db.cursor()
        cur.execute(
            """
            UPDATE watchlist_mt
            SET alert_threshold = %s, updated_at = NOW()
            WHERE organization_id = %s AND is_active = TRUE
        """,
            (threshold, str(current_user.organization_id)),
        )
        updated = cur.rowcount
        db.commit()

    return {
        "success": True,
        "message": f"{updated} items updated",
    }


async def import_watchlist_items_bulk(
    data,
    current_user,
    database_factory=None,
    watchlist_crud=None,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Bulk import manual watchlist items and return the created scan targets."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD

    with database_factory() as db:
        capacity = _get_watchlist_capacity(
            db,
            current_user,
            user_plan_getter=user_plan_getter,
            plan_limit_getter=plan_limit_getter,
        )
        remaining_slots = capacity["remaining_slots"]
        max_items = capacity["max_items"]
        current_count = capacity["current_count"]

        if remaining_slots == 0:
            _raise_watchlist_limit_exceeded(max_items, current_count)

        existing_app_nos = _load_existing_watchlist_application_numbers(
            db,
            current_user.organization_id,
        )

        created = 0
        failed = 0
        skipped = 0
        errors = []
        created_ids = []

        for index, item in enumerate(data.items):
            app_no_str = str(item.application_no).strip() if item.application_no else None
            if app_no_str and app_no_str in existing_app_nos:
                skipped += 1
                continue

            if created >= remaining_slots:
                errors.append(
                    {
                        "index": index,
                        "brand_name": item.brand_name,
                        "error": f"Izleme listesi limiti asildi ({max_items})",
                    }
                )
                failed += 1
                continue

            try:
                result = watchlist_crud.create(
                    db, current_user.organization_id, current_user.id, item
                )
                created += 1
                created_ids.append(UUID(str(result["id"])))
                if app_no_str:
                    existing_app_nos.add(app_no_str)
            except Exception as exc:
                failed += 1
                errors.append(
                    {
                        "index": index,
                        "brand_name": item.brand_name,
                        "error": str(exc),
                    }
                )

    return {
        "result": {
            "total": len(data.items),
            "created": created,
            "failed": failed,
            "skipped": skipped,
            "errors": errors,
        },
        "scan_item_ids": created_ids,
    }


async def preview_watchlist_portfolio_import(
    data,
    current_user,
    database_factory=None,
):
    """Preview duplicates before importing a holder or attorney portfolio into the watchlist."""
    if database_factory is None:
        database_factory = Database

    _require_portfolio_lookup_params(data)

    with database_factory() as db:
        cur = db.cursor()

        if data.holder_id:
            holder_id_str = str(data.holder_id)
            try:
                UUID(holder_id_str)
                cur.execute(
                    "SELECT application_no FROM trademarks WHERE holder_tpe_client_id = %s OR holder_id = %s",
                    (holder_id_str, holder_id_str),
                )
            except ValueError:
                cur.execute(
                    "SELECT application_no FROM trademarks WHERE holder_tpe_client_id = %s",
                    (holder_id_str,),
                )
        else:
            cur.execute(
                "SELECT application_no FROM trademarks WHERE attorney_no = %s OR attorney_tpe_client_id = %s",
                (str(data.attorney_no), str(data.attorney_no)),
            )

        rows = cur.fetchall()
        source_app_nos = {
            str(row["application_no"]).strip()
            for row in rows
            if row.get("application_no")
        }
        total_items = len(source_app_nos)

        if total_items == 0:
            return {
                "total_items": 0,
                "duplicate_count": 0,
                "can_add": 0,
            }

        existing_app_nos = _load_existing_watchlist_application_numbers(
            db,
            current_user.organization_id,
        )
        duplicate_count = len(source_app_nos.intersection(existing_app_nos))
        can_add = total_items - duplicate_count

    return {
        "total_items": total_items,
        "duplicate_count": duplicate_count,
        "can_add": can_add,
    }


async def import_watchlist_items_from_portfolio(
    data,
    current_user,
    database_factory=None,
    watchlist_crud=None,
    user_plan_getter=None,
    plan_limit_getter=None,
    watchlist_item_create_cls=None,
):
    """Bulk import watchlist items from a holder or attorney portfolio."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD
    if watchlist_item_create_cls is None:
        watchlist_item_create_cls = WatchlistItemCreate

    _require_portfolio_lookup_params(data)

    with database_factory() as db_perm:
        _ensure_portfolio_access_allowed(
            db_perm,
            current_user.id,
            user_plan_getter=user_plan_getter,
            plan_limit_getter=plan_limit_getter,
        )

    with database_factory() as db:
        cur = db.cursor()

        if data.holder_id:
            where_col = "holder_tpe_client_id"
            param = data.holder_id
        else:
            where_col = "attorney_no"
            param = data.attorney_no

        cur.execute(
            psql.SQL(
                """
            SELECT application_no, name, nice_class_numbers, image_path,
                   image_embedding::text, dinov2_embedding::text,
                   color_histogram::text, logo_ocr_text, text_embedding::text
            FROM trademarks
            WHERE {} = %s
            ORDER BY application_date DESC NULLS LAST
        """
            ).format(psql.Identifier(where_col)),
            (param,),
        )
        rows = cur.fetchall()

        if not rows:
            return {
                "result": {
                    "total": 0,
                    "created": 0,
                    "failed": 0,
                    "skipped": 0,
                    "errors": [],
                },
                "scan_item_ids": [],
            }

        capacity = _get_watchlist_capacity(
            db,
            current_user,
            user_plan_getter=user_plan_getter,
            plan_limit_getter=plan_limit_getter,
        )
        remaining_slots = capacity["remaining_slots"]
        max_items = capacity["max_items"]
        current_count = capacity["current_count"]
        existing_app_nos = _load_existing_watchlist_application_numbers(
            db,
            current_user.organization_id,
        )

        created = 0
        failed = 0
        skipped = 0
        errors = []
        created_ids = []
        limit_reached = False

        for index, trademark in enumerate(rows):
            app_no = trademark.get("application_no")
            app_no_str = str(app_no).strip() if app_no else None
            if app_no_str and app_no_str in existing_app_nos:
                skipped += 1
                continue

            if created >= remaining_slots:
                limit_reached = True
                break

            try:
                brand = trademark.get("name") or trademark.get("application_no") or "Unknown"
                classes = trademark.get("nice_class_numbers") or []
                classes = [nice_class for nice_class in classes if 1 <= nice_class <= 45]
                if not classes:
                    classes = [1]

                item_data = watchlist_item_create_cls(
                    brand_name=brand,
                    nice_class_numbers=classes,
                    application_no=trademark.get("application_no"),
                    similarity_threshold=data.similarity_threshold,
                )

                cur.execute("SAVEPOINT sp_bulk")
                result = watchlist_crud.create_with_embeddings(
                    db,
                    current_user.organization_id,
                    current_user.id,
                    item_data,
                    logo_path=trademark.get("image_path") or None,
                    logo_embedding=_parse_embedding_vector(trademark.get("image_embedding")),
                    logo_dinov2_embedding=_parse_embedding_vector(
                        trademark.get("dinov2_embedding")
                    ),
                    logo_color_histogram=_parse_embedding_vector(
                        trademark.get("color_histogram")
                    ),
                    logo_ocr_text=trademark.get("logo_ocr_text"),
                    text_embedding=_parse_embedding_vector(trademark.get("text_embedding")),
                    auto_commit=False,
                )
                cur.execute("RELEASE SAVEPOINT sp_bulk")
                created += 1
                created_ids.append(UUID(str(result["id"])))
                if app_no_str:
                    existing_app_nos.add(app_no_str)
            except Exception as exc:
                try:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_bulk")
                except Exception:
                    pass
                failed += 1
                errors.append(
                    {
                        "index": index,
                        "brand_name": trademark.get("name", ""),
                        "error": str(exc),
                    }
                )

        db.commit()

    return {
        "result": {
            "total": len(rows),
            "created": created,
            "failed": failed,
            "skipped": skipped,
            "errors": errors,
            "limit_reached": limit_reached,
            "max_allowed": max_items,
            "current_count": current_count + created,
        },
        "scan_item_ids": created_ids,
    }


def build_watchlist_upload_template():
    """Build the default watchlist upload template workbook in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Marka Listesi"

    headers = [
        ("Marka AdÄ± *", True),
        ("BaÅŸvuru No *", True),
        ("SÄ±nÄ±flar *", True),
        ("BÃ¼lten No", False),
    ]

    required_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    optional_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, (header, is_required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = required_fill if is_required else optional_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sample_data = [
        ["Ã–RNEK MARKA 1", "2023/12345", "9, 35", "305"],
        ["Ã–RNEK MARKA 2", "2023/67890", "25, 35, 42", "306"],
        ["Ã–RNEK MARKA 3", "2022/11111", "30, 43", ""],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.cell(row=6, column=1, value="* Zorunlu sÃ¼tunlar. BÃ¼lten No opsiyoneldir.")
    ws.cell(row=6, column=1).font = Font(italic=True, color="666666")

    ws.cell(row=7, column=1, value="SÄ±nÄ±flar: VirgÃ¼lle ayÄ±rarak yazÄ±n (Ã¶rn: 9, 35, 42)")
    ws.cell(row=7, column=1).font = Font(italic=True, color="666666")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def detect_watchlist_upload_columns(
    contents,
    filename,
    brand_name_variants,
    application_no_variants,
    class_variants,
    bulletin_variants,
    find_column=None,
    read_excel=None,
    read_csv=None,
):
    """Inspect an uploaded file and return its columns, samples, and auto-mappings."""
    if find_column is None:
        find_column = _find_column_match
    if read_excel is None:
        read_excel = pd.read_excel
    if read_csv is None:
        read_csv = pd.read_csv

    normalized_filename = filename.lower() if filename else ""

    try:
        if normalized_filename.endswith((".xlsx", ".xls")):
            df = read_excel(io.BytesIO(contents), nrows=5)
            df_count = read_excel(io.BytesIO(contents), usecols=[0])
            total_rows = len(df_count)
        elif normalized_filename.endswith(".csv"):
            df = read_csv(io.BytesIO(contents), nrows=5)
            df_count = read_csv(io.BytesIO(contents), usecols=[0])
            total_rows = len(df_count)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Desteklenmeyen dosya formati. Excel veya CSV yukleyin.",
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Dosya okunamadi: {str(exc)}",
        ) from exc

    original_columns = list(df.columns)
    normalized_columns = [str(col).lower().strip() for col in df.columns]

    auto_mappings = {
        "brand_name": find_column(normalized_columns, brand_name_variants),
        "application_no": find_column(normalized_columns, application_no_variants),
        "nice_classes": find_column(normalized_columns, class_variants),
        "bulletin_no": find_column(normalized_columns, bulletin_variants),
    }

    norm_to_orig = {str(col).lower().strip(): str(col) for col in original_columns}
    auto_mappings_orig = {
        "brand_name": norm_to_orig.get(auto_mappings["brand_name"])
        if auto_mappings["brand_name"]
        else None,
        "application_no": norm_to_orig.get(auto_mappings["application_no"])
        if auto_mappings["application_no"]
        else None,
        "nice_classes": norm_to_orig.get(auto_mappings["nice_classes"])
        if auto_mappings["nice_classes"]
        else None,
        "bulletin_no": norm_to_orig.get(auto_mappings["bulletin_no"])
        if auto_mappings["bulletin_no"]
        else None,
    }

    df.columns = original_columns
    sample_data = df.head(3).fillna("").to_dict("records")
    sample_data = [
        {key: str(value) if value != "" else "" for key, value in row.items()}
        for row in sample_data
    ]

    return {
        "columns": original_columns,
        "sample_data": sample_data,
        "auto_mappings": auto_mappings_orig,
        "total_rows": total_rows,
    }


async def store_watchlist_logo_upload(
    item_id,
    current_user,
    logo_filename,
    content_type,
    contents,
    database_factory=None,
    watchlist_crud=None,
    logos_dir=None,
    make_dirs=None,
    write_bytes=None,
    user_plan_getter=None,
    plan_limit_getter=None,
):
    """Persist an uploaded watchlist logo and return its stored path."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD
    if logos_dir is None:
        logos_dir = WATCHLIST_LOGOS_DIR
    if make_dirs is None:
        make_dirs = os.makedirs
    if write_bytes is None:
        def write_bytes(path, payload):
            with open(path, "wb") as file_obj:
                file_obj.write(payload)

    with database_factory() as db:
        item = watchlist_crud.get_by_id(db, item_id, current_user.organization_id)
        if not item:
            _raise_watchlist_item_not_found()
        _ensure_logo_tracking_allowed(
            db,
            current_user.id,
            user_plan_getter=user_plan_getter,
            plan_limit_getter=plan_limit_getter,
        )

    if not content_type or not content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Dosya bir gorsel olmali (PNG, JPG, WEBP)")

    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya boyutu 5MB'yi asamaz")

    org_dir = os.path.join(logos_dir, str(current_user.organization_id))
    make_dirs(org_dir, exist_ok=True)

    ext = os.path.splitext(logo_filename or "logo.png")[1] or ".png"
    filename = f"{item_id}{ext}"
    filepath = os.path.join(org_dir, filename)
    write_bytes(filepath, contents)

    with database_factory() as db:
        watchlist_crud.update_logo(db, item_id, logo_path=filepath)

    return {
        "success": True,
        "message": "Logo yuklendi, embeddingler olusturuluyor...",
        "item_id": item_id,
        "filepath": filepath,
    }


def _image_media_type_for_path(path):
    """Resolve a file extension into an image media type."""
    ext = os.path.splitext(path)[1].lower()
    return {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
    }.get(ext, "image/png")


async def resolve_watchlist_logo_file(
    item_id,
    database_factory=None,
    file_exists=None,
    project_root=None,
):
    """Resolve the best on-disk file for a watchlist logo image request."""
    if database_factory is None:
        database_factory = Database
    if file_exists is None:
        file_exists = os.path.isfile
    if project_root is None:
        project_root = PROJECT_ROOT
    project_root = os.fspath(project_root)

    with database_factory() as db:
        cur = db.cursor()
        cur.execute("SELECT logo_path FROM watchlist_mt WHERE id = %s", (str(item_id),))
        row = cur.fetchone()
        if not row:
            _raise_watchlist_item_not_found()

    logo_path = row.get("logo_path") if row else None
    if not logo_path:
        raise HTTPException(status_code=404, detail="Logo bulunamadi")
    if ".." in logo_path:
        raise HTTPException(status_code=400, detail="Invalid path")

    if file_exists(logo_path):
        return {"path": logo_path, "media_type": _image_media_type_for_path(logo_path)}

    full_path = os.path.join(project_root, logo_path.replace("/", os.sep))
    if file_exists(full_path):
        return {"path": full_path, "media_type": _image_media_type_for_path(full_path)}

    import re as _re

    rel_match = _re.search(r"((?:bulletins|uploads)[/\\].+)", logo_path.replace("\\", "/"))
    if rel_match:
        relative = rel_match.group(1)
        candidate = os.path.join(project_root, relative.replace("/", os.sep))
        if file_exists(candidate):
            return {"path": candidate, "media_type": _image_media_type_for_path(candidate)}

    raise HTTPException(status_code=404, detail="Logo bulunamadi")


async def delete_watchlist_logo_asset(
    item_id,
    current_user,
    database_factory=None,
    watchlist_crud=None,
    file_exists=None,
    remove_file=None,
):
    """Delete a watchlist logo file and clear its stored embeddings."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD
    if file_exists is None:
        file_exists = os.path.isfile
    if remove_file is None:
        remove_file = os.remove

    with database_factory() as db:
        item = watchlist_crud.get_by_id(db, item_id, current_user.organization_id)
        if not item:
            _raise_watchlist_item_not_found()

    logo_path = item.get("logo_path")
    if logo_path and file_exists(logo_path):
        try:
            remove_file(logo_path)
        except OSError:
            pass

    with database_factory() as db:
        watchlist_crud.clear_logo(db, item_id)

    return {
        "success": True,
        "message": "Logo silindi",
    }


def process_watchlist_logo_embeddings(
    item_id,
    filepath,
    logger=None,
    database_factory=None,
    watchlist_crud=None,
    embedding_generator=None,
    traceback_formatter=None,
):
    """Generate and store visual embeddings for a watchlist logo file."""
    if database_factory is None:
        database_factory = Database
    if watchlist_crud is None:
        watchlist_crud = WatchlistCRUD
    if embedding_generator is None:
        from watchlist.scanner import generate_logo_embeddings

        embedding_generator = generate_logo_embeddings
    if traceback_formatter is None:
        import traceback as _traceback

        traceback_formatter = _traceback.format_exc

    if logger:
        logger.info(f"[LOGO] Generating embeddings for watchlist {item_id}")

    try:
        result = embedding_generator(filepath)
        if not result:
            if logger:
                logger.warning(f"[LOGO] No embeddings generated for {item_id}")
            return

        with database_factory() as db:
            watchlist_crud.update_logo(
                db,
                item_id,
                logo_path=filepath,
                logo_embedding=result.get("clip_embedding"),
                dino_embedding=result.get("dino_embedding"),
                color_histogram=result.get("color_histogram"),
                logo_ocr_text=result.get("ocr_text"),
            )

        if logger:
            logger.info(f"[LOGO] Embeddings stored for {item_id}")
    except Exception as exc:
        if logger:
            logger.error(f"[LOGO] Failed for {item_id}: {exc}")
            logger.error(traceback_formatter())


def _parse_watchlist_nice_classes(value):
    """Parse Nice class values from uploads into a deduplicated list of integers."""
    if pd.isna(value) or not value:
        return []

    numbers = re.findall(r"\d+", str(value))
    classes = []
    for number in numbers:
        parsed = int(number)
        if 1 <= parsed <= 45:
            classes.append(parsed)

    return sorted(list(set(classes)))


def _read_watchlist_upload_dataframe(
    contents,
    filename,
    unsupported_detail,
    parse_error_builder,
    read_excel=None,
    read_csv=None,
):
    """Load an upload file into a dataframe using the endpoint-specific error shape."""
    if read_excel is None:
        read_excel = pd.read_excel
    if read_csv is None:
        read_csv = pd.read_csv

    normalized_filename = filename.lower() if filename else ""

    try:
        if normalized_filename.endswith((".xlsx", ".xls")):
            return read_excel(io.BytesIO(contents))
        if normalized_filename.endswith(".csv"):
            return read_csv(io.BytesIO(contents))
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=unsupported_detail,
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=parse_error_builder(exc),
        ) from exc


def _build_watchlist_file_upload_result(
    total_rows,
    added_count,
    skipped_count,
    error_count,
    warnings,
    skipped_items,
    error_items,
):
    """Build the shared file-upload response payload."""
    message_parts = [f"{added_count} marka eklendi"]
    if skipped_count > 0:
        message_parts.append(f"{skipped_count} zaten mevcut (atlandi)")
    if error_count > 0:
        message_parts.append(f"{error_count} hatali satir")

    return FileUploadResult(
        success=True,
        message=", ".join(message_parts),
        summary=FileUploadSummary(
            total_rows=total_rows,
            added=added_count,
            skipped=skipped_count,
            errors=error_count,
        ),
        warnings=warnings,
        skipped_items=skipped_items[:10],
        error_items=error_items[:10],
    )


def _import_watchlist_upload_rows(
    df,
    current_user,
    brand_column,
    application_column,
    class_column,
    bulletin_column,
    warnings,
    require_application_number,
    generate_missing_application_number,
    require_classes,
    database_factory=None,
    user_plan_getter=None,
    plan_limit_getter=None,
    parse_nice_classes=None,
    uuid_factory=None,
):
    """Persist parsed upload rows and return the response payload plus scan targets."""
    if database_factory is None:
        database_factory = Database
    if parse_nice_classes is None:
        parse_nice_classes = _parse_watchlist_nice_classes
    if uuid_factory is None:
        uuid_factory = uuid4

    with database_factory() as db:
        capacity = _get_watchlist_capacity(
            db,
            current_user,
            user_plan_getter=user_plan_getter,
            plan_limit_getter=plan_limit_getter,
        )
        remaining_slots = capacity["remaining_slots"]
        max_items = capacity["max_items"]
        existing_app_nos = {
            str(value).strip().lower()
            for value in _load_existing_watchlist_application_numbers(
                db,
                current_user.organization_id,
            )
            if value
        }

        added_count = 0
        skipped_count = 0
        error_count = 0
        skipped_items = []
        error_items = []
        created_ids = []

        cur = db.cursor()
        org_id = str(current_user.organization_id)
        user_id = str(current_user.id)

        for idx, row in df.iterrows():
            row_num = idx + 2
            brand_name = None

            try:
                if added_count >= remaining_slots:
                    error_count += 1
                    error_items.append(
                        FileUploadErrorItem(
                            row=row_num,
                            error=f"Izleme listesi limiti asildi ({max_items})",
                        )
                    )
                    continue

                brand_name = str(row.get(brand_column, "")).strip()
                if not brand_name or brand_name.lower() in ["nan", "none", ""]:
                    error_count += 1
                    error_items.append(
                        FileUploadErrorItem(
                            row=row_num,
                            error="Marka adi bos",
                        )
                    )
                    continue

                application_value = row.get(application_column, "") if application_column else ""
                app_no = str(application_value).strip()
                if not app_no or app_no.lower() in ["nan", "none", ""]:
                    if generate_missing_application_number:
                        app_no = f"WL-{uuid_factory().hex[:8].upper()}"
                    elif require_application_number:
                        error_count += 1
                        error_items.append(
                            FileUploadErrorItem(
                                row=row_num,
                                brand_name=brand_name,
                                error="Basvuru numarasi bos",
                            )
                        )
                        continue

                classes_raw = row.get(class_column, "") if class_column and class_column in row else ""
                classes_str = str(classes_raw).strip() if classes_raw is not None else ""
                if classes_str and classes_str.lower() not in ["nan", "none", ""]:
                    nice_classes = parse_nice_classes(classes_raw)
                else:
                    nice_classes = []

                if require_classes and not nice_classes:
                    error_count += 1
                    error_items.append(
                        FileUploadErrorItem(
                            row=row_num,
                            brand_name=brand_name,
                            error="Sinif bilgisi bos veya gecersiz",
                        )
                    )
                    continue

                bulletin_no = None
                if bulletin_column and bulletin_column in row:
                    bulletin_no = str(row.get(bulletin_column, "")).strip()
                    if bulletin_no.lower() in ["nan", "none", ""]:
                        bulletin_no = None

                app_no_key = app_no.lower()
                if app_no_key in existing_app_nos:
                    skipped_count += 1
                    skipped_items.append(
                        FileUploadSkippedItem(
                            row=row_num,
                            brand_name=brand_name,
                            application_no=app_no,
                            reason="Zaten mevcut",
                        )
                    )
                    continue

                item_id = uuid_factory()
                cur.execute(
                    """
                    INSERT INTO watchlist_mt (
                        id, organization_id, user_id, brand_name,
                        nice_class_numbers, customer_application_no, customer_bulletin_no,
                        alert_threshold, is_active, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s,
                        0.7, TRUE, NOW(), NOW()
                    )
                """,
                    (
                        str(item_id),
                        org_id,
                        user_id,
                        brand_name,
                        nice_classes,
                        app_no,
                        bulletin_no,
                    ),
                )

                added_count += 1
                existing_app_nos.add(app_no_key)
                created_ids.append(item_id)
            except Exception as exc:
                error_count += 1
                error_items.append(
                    FileUploadErrorItem(
                        row=row_num,
                        brand_name=brand_name,
                        error=str(exc)[:100],
                    )
                )

        db.commit()

    return {
        "result": _build_watchlist_file_upload_result(
            total_rows=len(df),
            added_count=added_count,
            skipped_count=skipped_count,
            error_count=error_count,
            warnings=warnings,
            skipped_items=skipped_items,
            error_items=error_items,
        ),
        "scan_item_ids": created_ids,
    }


async def import_watchlist_upload_with_mapping(
    contents,
    filename,
    column_mapping,
    current_user,
    database_factory=None,
    user_plan_getter=None,
    plan_limit_getter=None,
    read_excel=None,
    read_csv=None,
    parse_nice_classes=None,
    uuid_factory=None,
):
    """Import a watchlist file using an explicit client-provided column mapping."""
    try:
        mappings = json.loads(column_mapping)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Gecersiz sutun eslestirme formati",
        ) from exc

    if "nice_class_numbers" in mappings and "nice_classes" not in mappings:
        mappings["nice_classes"] = mappings.pop("nice_class_numbers")

    if not mappings.get("brand_name"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Eksik zorunlu eslestirme: brand_name",
        )

    df = _read_watchlist_upload_dataframe(
        contents=contents,
        filename=filename,
        unsupported_detail="Desteklenmeyen dosya formati",
        parse_error_builder=lambda exc: f"Dosya okunamadi: {str(exc)}",
        read_excel=read_excel,
        read_csv=read_csv,
    )

    if df.empty:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Dosya bos",
        )

    rename_map = {value: key for key, value in mappings.items() if value}
    df = df.rename(columns=rename_map)
    df.columns = [str(col).lower().strip() for col in df.columns]

    warnings = []
    if not mappings.get("bulletin_no"):
        warnings.append(
            FileUploadWarning(
                column="Bulten No",
                message="Bulten numarasi sutunu eslestirme yapilmadi. Bu opsiyonel bir alandir.",
            )
        )

    return _import_watchlist_upload_rows(
        df=df,
        current_user=current_user,
        brand_column="brand_name",
        application_column="application_no",
        class_column="nice_classes" if "nice_classes" in df.columns else None,
        bulletin_column="bulletin_no" if "bulletin_no" in df.columns else None,
        warnings=warnings,
        require_application_number=False,
        generate_missing_application_number=True,
        require_classes=False,
        database_factory=database_factory,
        user_plan_getter=user_plan_getter,
        plan_limit_getter=plan_limit_getter,
        parse_nice_classes=parse_nice_classes,
        uuid_factory=uuid_factory,
    )


async def import_watchlist_upload_file(
    contents,
    filename,
    current_user,
    brand_name_variants,
    application_no_variants,
    class_variants,
    bulletin_variants,
    find_column=None,
    parse_nice_classes=None,
    database_factory=None,
    user_plan_getter=None,
    plan_limit_getter=None,
    read_excel=None,
    read_csv=None,
    uuid_factory=None,
):
    """Import a watchlist file using auto-detected mandatory columns."""
    if find_column is None:
        find_column = _find_column_match

    df = _read_watchlist_upload_dataframe(
        contents=contents,
        filename=filename,
        unsupported_detail={
            "error": "unsupported_format",
            "message": "Desteklenmeyen dosya formati",
            "detail": "Lutfen Excel (.xlsx, .xls) veya CSV (.csv) dosyasi yukleyin.",
        },
        parse_error_builder=lambda exc: {
            "error": "parse_error",
            "message": "Dosya okunamadi",
            "detail": str(exc),
        },
        read_excel=read_excel,
        read_csv=read_csv,
    )

    if df.empty:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error": "empty_file",
                "message": "Dosya bos",
            },
        )

    original_columns = list(df.columns)
    df.columns = [str(col).lower().strip() for col in df.columns]

    brand_col = find_column(df.columns.tolist(), brand_name_variants)
    app_no_col = find_column(df.columns.tolist(), application_no_variants)
    class_col = find_column(df.columns.tolist(), class_variants)
    bulletin_col = find_column(df.columns.tolist(), bulletin_variants)

    missing_columns = []
    if not brand_col:
        missing_columns.append(
            {
                "column": "Marka Adi",
                "variants": "marka adi, brand name, name, isim",
                "reason": "Hangi markalarin izlenecegini belirler",
            }
        )
    if not app_no_col:
        missing_columns.append(
            {
                "column": "Basvuru No",
                "variants": "basvuru no, application no, app no",
                "reason": "Mukerrer kontrol ve cakisma filtreleme icin gerekli",
            }
        )
    if not class_col:
        missing_columns.append(
            {
                "column": "Siniflar",
                "variants": "sinif, siniflar, nice class, classes",
                "reason": "Hangi siniflarda arama yapilacagini belirler",
            }
        )

    if missing_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error": "missing_mandatory_columns",
                "message": f"{len(missing_columns)} zorunlu sutun eksik",
                "missing_columns": missing_columns,
                "found_columns": original_columns,
                "required_columns": [
                    {"name": "Marka Adi", "variants": "marka adi, brand name, name"},
                    {"name": "Basvuru No", "variants": "basvuru no, application no"},
                    {"name": "Siniflar", "variants": "sinif, siniflar, nice class, classes"},
                ],
                "optional_columns": [
                    {"name": "Bulten No", "variants": "bulten no, bulletin no"},
                ],
                "example": {
                    "headers": ["Marka Adi", "Basvuru No", "Siniflar", "Bulten No"],
                    "rows": [
                        ["ORNEK MARKA", "2023/12345", "9, 35, 42", "305"],
                        ["DIGER MARKA", "2023/67890", "25, 35", "306"],
                    ],
                },
            },
        )

    warnings = []
    if not bulletin_col:
        warnings.append(
            FileUploadWarning(
                column="Bulten No",
                message="Bulten numarasi sutunu bulunamadi. Bu opsiyonel bir alandir.",
            )
        )

    return _import_watchlist_upload_rows(
        df=df,
        current_user=current_user,
        brand_column=brand_col,
        application_column=app_no_col,
        class_column=class_col,
        bulletin_column=bulletin_col,
        warnings=warnings,
        require_application_number=True,
        generate_missing_application_number=False,
        require_classes=True,
        database_factory=database_factory,
        user_plan_getter=user_plan_getter,
        plan_limit_getter=plan_limit_getter,
        parse_nice_classes=parse_nice_classes,
        uuid_factory=uuid_factory,
    )
